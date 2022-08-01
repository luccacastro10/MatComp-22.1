import math
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook


# Método dos Mínimos Quadrados: 
#--------------------------------------------
# Recebe uma matriz de pontos, cada um contendo px e py
# Realiza o cálculo dos mínimios quadrados e retorna o coeficiente angular e linear
# da melhor reta que se ajusta ao conjunto de dados
#--------------------------------------------
def leastSquares(points):

    n = len(points)
    
    xSum  = 0
    ySum  = 0
    x2Sum = 0
    xySum = 0

    for point in points:
        xSum  += point[0]
        ySum  += point[1]
        x2Sum += point[0]**2
        xySum += point[0]*point[1]

    m = ((n*xySum) - (xSum*ySum))/((n*x2Sum) - (xSum**2))
    n = (ySum - (m*xSum))/n

    return [m,n]


# Método da Regressão Linear Logística: 
#--------------------------------------------
# Recebe uma instância da classe municipio.y
# Reajusta o conjusto de dados para a forma linearizada da equação logística
# Aplica o método dos mínimos quadrados e avalia os parâmetros r e k em função dos
# coeficientes angular e linear encontrados
# Salva os parâmetros no atributo logisticParameters do objeto city
# Além disso, aplica o ajuste do ponto de inflexão da equação logística, salvando os
# resultados no atributo adjustedLogisticParameters do objeto city
#--------------------------------------------
def logisticRegression(city):

    k = 0
    r = 0

    t = []
    p = []
    p1 = []
    p2 = []

    for point in city.points:
        t.append(point[0])
        p.append(point[1])
    
    p1 = p[:(len(p)-1)]
    p2 = p[1:]
    p0 = p[0]

    firstRegression = []
    for i in range(len(p1)):
        point = []

        point.append(p1[i])

        #y = dp/dt * 1/p
        dpdt = (p2[i]-p1[i])/((t[i+1] - t[i])*p1[i])

        point.append(dpdt)

        firstRegression.append(point)

    firstResult = leastSquares(firstRegression) # Aplica mínimos quadrados na equação linearizada

    r = firstResult[1]
    k = (-1)*firstResult[1]/firstResult[0]

    city.logisticParameters = [p0,k,r] # Salva resultados da regressão linear logística

    k   = p[8] + p[9]                   # Cálculo de k e r via ajuste do ponto de inflexão 
    t_  = (t[8] + t[9])/2
    r = (1/t_)*math.log((k-p0)/p0)

    city.adjustedLogisticParameters = [p0,k,r] # Salva resultados da aproximação via ponto de inflexão


# Método da Descida de Gradiente: 
#--------------------------------------------
# Recebe uma instância da classe municipio.y
# Calcula gradiente da função de custo logístico
# Ajusta os parâmetros até que o stepSize seja tão pequeno quanto se queira
# Salva os parâmetros no atributo gradientParameters do objeto city
#--------------------------------------------
def gradientDescent(city):

    p0 = city.points[0][1]
    k = 1.15*city.points[-1][1] # Tomamos o ponto mais alto como valor de partida para a capacidade suporte

    r = city.adjustedLogisticParameters[2] # Tomamos os parâmetros obtidos pelo ponto de inflexão como nosso ponto de partida da descida de gradiente

    limitStepSize = 0.00001
    learningRateR = 0.000001
    learningRateK = 0.001

    iterations = 0
    while True:

        kGradient = 0
        rGradient = 0
        
        #Calculando os gradientes
        for point in city.points:

            exp = np.exp((-1)*r*point[0])
            exp2 = np.exp((-2)*r*point[0])

            denominador = (p0+exp*(k-p0))*(p0-k*p0+exp*(k-p0))

            x = point[0]
            y = point[1]
            
            kGradient += (k*(p0**2)-k*exp*(p0**2)-exp2*y*(p0**2)+2*exp*y*(p0**2)-y*(p0**2)+k*exp2*y*p0-k*exp*y*p0)/(k*denominador)
            rGradient += ((-1)*(k**2)*exp2*y*x - exp2*y*(p0**2)*x+exp*y*(p0**2)*x-k*exp*(p0**2)*x-k*exp*y*p0*x+2*k*exp2*y*p0*x+(k**2)*exp*p0*x)/(denominador)
        
        kGradient = (-2)*kGradient/len(city.points)
        rGradient = (-2)*rGradient/len(city.points)

        setpSize_k = kGradient * learningRateK
        setpSize_r = rGradient * learningRateR

        k = k - setpSize_k
        r = (r - setpSize_r)

        if ( ((abs(setpSize_k) < limitStepSize) and (abs(setpSize_r) < limitStepSize)) ):
            break
    
        iterations += 1 

        if (iterations%1000 == 0):
            print("StepSize K: " + str(setpSize_k) + " | StepSize R: " + str(setpSize_r), end="\r")

    city.gradientParameters = [p0,k,r]



# Método plotLogistic: 
#--------------------------------------------
# Recebe uma instância da classe municipio.y
# Plota os gráficos a partir dos diferentes valores de k e r encontrados
# Salva as imagens no diretório results/images/
#--------------------------------------------
def plotLogistic(city):

    figure, axis = plt.subplots(1, 3)

    methods = [city.logisticParameters, city.adjustedLogisticParameters, city.gradientParameters]
    methodsNames = ["Linear Regression: ", "Linear Regression + Adjust: ", "Gradient Descent: "]

    i = 0

    for i in range(3):

        p0 = methods[i][0]
        k = methods[i][1]
        r = methods[i][2]
        x = np.linspace(city.points[0][0], city.points[-1][0],1000)
        y = (p0*k)/((k-p0)*np.exp((-1)*r*x) + p0)


        if not len(city.points):
            city.loadPoints()

        x2 = []
        y2 = []

        for point in city.points:
            x2.append(point[0])
            y2.append(point[1])

        axis[i].set_title(methodsNames[i] + city.name)
        axis[i].scatter(x2,y2)
        axis[i].plot(x,y, 'g')


        i += 1

    plt.gcf().set_size_inches(16, 5)
    plt.savefig(fname="results/images/logisticRegression_" + city.name + ".jpg")


# Método saveExcel: 
#--------------------------------------------
# Recebe uma instância da classe municipio.y
# Armazena dados dos erros de cada método num arquivo excel
#--------------------------------------------
def saveExcel(city):
  
    workbook = Workbook()
    
    sheet = workbook.active 

    c1 = sheet.cell(row = 1, column = 1)   
    c1.value = "Ano"
    
    c2 = sheet.cell(row= 1 , column = 2) 
    c2.value = "Valor Real"

    c3 = sheet.cell(row= 1 , column = 3) 
    c3.value = "Logístico"

    c4 = sheet.cell(row= 1 , column = 4) 
    c4.value = "Logístico Ajustado"

    c5 = sheet.cell(row= 1 , column = 5) 
    c5.value = "Descida de Gradiente"

    c6 = sheet.cell(row= 1 , column = 6) 
    c6.value = "Erro Logístico"

    c7 = sheet.cell(row= 1 , column = 7) 
    c7.value = "Erro Logístico Ajustado"

    c8 = sheet.cell(row= 1 , column = 8) 
    c8.value = "Erro Descida de Gradiente"

    for i in range(len(city.points)):
        c = sheet.cell(row = 2 + i, column = 1)
        c.value = city.points[i][0]

    for i in range(len(city.points)):
        c = sheet.cell(row = 2 + i, column = 2)
        c.value = city.points[i][1]

    for i in range(len(city.logisticResult)):
        c = sheet.cell(row = 2 + i, column = 3)
        c.value = city.logisticResult[i]

    for i in range(len(city.logisticResult)):
        c = sheet.cell(row = 2 + i, column = 4)
        c.value = city.adjustedLogisticResult[i]

    for i in range(len(city.logisticResult)):
        c = sheet.cell(row = 2 + i, column = 5)
        c.value = city.gradientResult[i]

    for i in range(len(city.logisticError)):
        c = sheet.cell(row = 2 + i, column = 6)
        c.value = city.logisticError[i]

    for i in range(len(city.logisticError)):
        c = sheet.cell(row = 2 + i, column = 7)
        c.value = city.adjustedLogisticError[i]

    for i in range(len(city.logisticError)):
        c = sheet.cell(row = 2 + i, column = 8)
        c.value = city.gradientError[i]


    workbook.save(filename="results/error_results/tables/logisticRegression_" + city.name + ".xlsx")

# Método getError: 
#--------------------------------------------
# Recebe uma instância da classe municipio.y
# Calcula os erros de cada um dos métodos de ajuste
# Salva gráficos de erros no diretório results/error_results/graphs/
# Salva tabels de erros no diretório results/error_results/tables/
#--------------------------------------------
def getError(city):
    plt.close()

    parametersList = [city.logisticParameters, city.adjustedLogisticParameters, city.gradientParameters]
    resultList     = [city.logisticResult, city.adjustedLogisticResult, city.gradientResult] 
    errorList      = [city.logisticError, city.adjustedLogisticError, city.gradientError]
    colors = ["blue", "green", "purple"]
    labels = ["Ajuste Linear", "Ajuste Linear Ajustado", "Descida de Gradiente"]

    for j in range(len(parametersList)):
        p0 = parametersList[j][0]
        k = parametersList[j][1]
        r = parametersList[j][2]
        _t = []

        for i in range(len(city.points)):
            t = city.points[i][0]
            _t.append(t)
            realValue = city.points[i][1]

            y = (p0*k)/((k-p0)*np.exp((-1)*r*city.points[i][0]) + p0)

            error = abs( y - realValue)/realValue

            resultList[j].append(y)
            errorList[j].append(error)

        plt.scatter(_t, errorList[j], marker='o', color='red')
        plt.plot(_t, errorList[j], color=colors[j],markerfacecolor='red',
                markersize=10, label=labels[j])
        plt.grid(True)
        plt.title('Distribuição de Erros - ' + city.name)
        plt.legend()
    plt.savefig(fname="results/error_results/graphs/erros_" + city.name + ".jpg")

    saveExcel(city)
